# requirements.txt:
# fastapi
# uvicorn
# python-dotenv
# openai
# httpx
# openpyxl
# PyMuPDF

import json
import os
import re
import ssl
import tempfile
from pathlib import Path
from typing import Any

import fitz
import httpx
import openpyxl
import uvicorn
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openai import AsyncOpenAI
from pydantic import BaseModel

load_dotenv(Path(".") / ".env")

LITELLM_API_BASE = os.getenv("LITELLM_API_BASE", "")
LITELLM_API_KEY = os.getenv("LITELLM_API_KEY", "")
LITELLM_SSL_CERT = os.getenv("LITELLM_SSL_CERT", "")
LITELLM_TIMEOUT = int(os.getenv("LITELLM_TIMEOUT", "120"))
LITELLM_MODEL = os.getenv("LITELLM_MODEL", "claude-sonnet-4-6")
API_PORT = int(os.getenv("API_PORT", "8080"))


def _make_ssl_context() -> ssl.SSLContext | bool:
    if not LITELLM_SSL_CERT:
        return True
    ctx = ssl.create_default_context()
    cert_content = LITELLM_SSL_CERT.replace("\\n", "\n")
    ctx.load_verify_locations(cadata=cert_content)
    return ctx


http_client = httpx.AsyncClient(
    verify=_make_ssl_context(),
    timeout=httpx.Timeout(connect=30.0, read=float(LITELLM_TIMEOUT), write=60.0, pool=30.0),
)
base_url = LITELLM_API_BASE.rstrip("/") + "/v1" if LITELLM_API_BASE else ""
client = AsyncOpenAI(
    api_key=LITELLM_API_KEY,
    base_url=base_url,
    http_client=http_client,
    timeout=LITELLM_TIMEOUT,
)

app = FastAPI(title="DataChat API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:*", "http://127.0.0.1:*"],
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1)(:\d+)?",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- Models ---

class ContextPayload(BaseModel):
    type: str = "custom"
    schema_: list[dict[str, Any]] | None = None
    sampleRows: list[dict[str, Any]] | None = None
    text: str | None = None
    metadata: dict[str, Any] | None = None

    class Config:
        fields = {"schema_": "schema"}
        populate_by_name = True


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    context: ContextPayload
    messages: list[ChatMessage]
    responseFormat: str = "auto"


class JsonIngestRequest(BaseModel):
    data: list[dict[str, Any]]
    metadata: dict[str, Any] | None = None


# --- Helpers ---

def build_system_prompt(context: ContextPayload, response_format: str) -> str:
    if context.type == "structured":
        schema_str = json.dumps(context.schema_, indent=2) if context.schema_ else "[]"
        sample_str = json.dumps(context.sampleRows[:5], indent=2) if context.sampleRows else "[]"
        total = context.metadata.get("totalRows", "unknown") if context.metadata else "unknown"
        prompt = (
            f"You are an expert data analyst. You have access to a dataset with {total} rows.\n\n"
            f"Schema (column names and types):\n{schema_str}\n\n"
            f"Sample rows (first 5):\n{sample_str}\n\n"
            "You can answer questions about this data: summarize, filter, aggregate, compare, "
            "find trends, calculate statistics, or suggest visualizations. "
            "Be precise with numbers. When computing aggregates, describe your methodology."
        )
    elif context.type == "document":
        meta_str = json.dumps(context.metadata) if context.metadata else ""
        text_preview = (context.text or "")[:8000]
        prompt = (
            f"You are a document analyst. {f'Document metadata: {meta_str}. ' if meta_str else ''}"
            f"Document content:\n{text_preview}\n\n"
            "Answer questions about this document. Quote relevant passages when possible. "
            "If the document is long, reference page numbers or sections."
        )
    else:
        prompt = (
            f"You are an AI assistant. Context provided:\n{context.text or ''}\n\n"
            "Answer questions based on this context. Be helpful and precise."
        )

    if response_format == "chart":
        prompt += (
            "\n\nIMPORTANT: The user wants a chart. Respond with ONLY valid JSON in this format: "
            '{"type": "bar"|"line"|"donut"|"scatter", "title": "Chart Title", '
            '"data": [{...}], "xKey": "fieldName", "yKeys": ["fieldName1"]}. '
            "The data array must contain the actual computed values for the chart."
        )
    elif response_format == "table":
        prompt += (
            "\n\nIMPORTANT: The user wants tabular output. Respond with ONLY a JSON array of objects, "
            "e.g. [{col1: val, col2: val}, ...]. Include relevant columns and computed values."
        )
    elif response_format == "auto":
        prompt += (
            "\n\nResponse format rules:\n"
            "- If the answer is best shown as a chart, wrap it in ```chart\\n{json}\\n``` "
            "where json is: {type, title, data, xKey, yKeys}\n"
            "- If best as a table, wrap it in ```table\\n[{...}, ...]\\n```\n"
            "- Otherwise respond in plain text with markdown formatting.\n"
            "- For simple text answers, just respond normally."
        )
    return prompt


def parse_response(text: str) -> dict[str, Any]:
    chart_match = re.search(r"```chart\s*\n([\s\S]*?)\n```", text)
    if chart_match:
        try:
            data = json.loads(chart_match.group(1))
            return {"response": text, "type": "chart", "data": data}
        except json.JSONDecodeError:
            pass

    table_match = re.search(r"```table\s*\n([\s\S]*?)\n```", text)
    if table_match:
        try:
            data = json.loads(table_match.group(1))
            return {"response": text, "type": "table", "data": data}
        except json.JSONDecodeError:
            pass

    try:
        data = json.loads(text)
        if isinstance(data, dict) and "type" in data:
            return {"response": text, "type": "chart", "data": data}
        if isinstance(data, list):
            return {"response": text, "type": "table", "data": data}
    except json.JSONDecodeError:
        pass

    return {"response": text, "type": "text", "data": None}


def detect_column_type(values: list) -> str:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "text"

    numeric_count = 0
    for v in non_empty:
        try:
            float(v)
            numeric_count += 1
        except (ValueError, TypeError):
            break
    if numeric_count == len(non_empty):
        return "numeric"

    date_pattern = re.compile(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}")
    if all(date_pattern.search(str(v)) for v in non_empty[:20]):
        return "date"

    unique = set(str(v) for v in non_empty)
    if len(unique) < 20:
        return "categorical"

    return "text"


def infer_schema(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    if not rows:
        return []
    columns = list(rows[0].keys())
    schema = []
    for col in columns:
        values = [row.get(col) for row in rows[:100]]
        schema.append({"name": col, "type": detect_column_type(values)})
    return schema


# --- Endpoints ---

@app.post("/api/chat")
async def chat(request: ChatRequest):
    import asyncio
    global client, http_client

    if not LITELLM_API_BASE:
        raise HTTPException(status_code=503, detail="LLM not configured — check .env credentials")

    try:
        system_prompt = build_system_prompt(request.context, request.responseFormat)
        messages = [{"role": "system", "content": system_prompt}]
        recent = request.messages[-20:] if len(request.messages) > 20 else request.messages
        messages.extend([{"role": m.role, "content": m.content} for m in recent])

        last_err = None
        for _attempt in range(3):
            try:
                completion = await client.chat.completions.create(
                    model=LITELLM_MODEL,
                    messages=messages,
                    temperature=0.3,
                )
                text = completion.choices[0].message.content
                return parse_response(text)
            except Exception as attempt_err:
                last_err = attempt_err
                print(f"[chat] attempt {_attempt+1} failed: {type(attempt_err).__name__}: {attempt_err}", flush=True)
                if _attempt < 2:
                    try:
                        await http_client.aclose()
                    except Exception:
                        pass
                    http_client = httpx.AsyncClient(
                        verify=_make_ssl_context(),
                        timeout=httpx.Timeout(connect=30.0, read=float(LITELLM_TIMEOUT), write=60.0, pool=30.0),
                    )
                    client = AsyncOpenAI(
                        api_key=LITELLM_API_KEY,
                        base_url=base_url,
                        http_client=http_client,
                        timeout=LITELLM_TIMEOUT,
                    )
                    await asyncio.sleep(2 * (_attempt + 1))
        raise HTTPException(status_code=500, detail=f"LLM call failed after 3 attempts: {last_err}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ingest/excel")
async def ingest_excel(file: UploadFile = File(...)):
    try:
        suffix = Path(file.filename or "upload.xlsx").suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw:
                continue
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows_raw[0])]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:]]
            schema = infer_schema(data_rows[:100])
            sheets.append({
                "name": sheet_name,
                "schema": schema,
                "rows": data_rows[:100],
                "totalRows": len(data_rows),
                "sampleRows": data_rows[:10],
            })
        wb.close()
        os.unlink(tmp_path)
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ingest/pdf")
async def ingest_pdf(file: UploadFile = File(...)):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        doc = fitz.open(tmp_path)
        chunks = []
        for page in doc:
            chunks.append(page.get_text())
        full_text = "\n".join(chunks)
        title = doc.metadata.get("title", "") if doc.metadata else ""
        page_count = len(doc)
        doc.close()
        os.unlink(tmp_path)

        return {
            "pages": page_count,
            "text": full_text,
            "chunks": chunks,
            "metadata": {"title": title, "pages": page_count},
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ingest/json")
async def ingest_json(request: JsonIngestRequest):
    try:
        rows = request.data
        schema = infer_schema(rows)
        return {
            "schema": schema,
            "rows": rows[:100],
            "totalRows": len(rows),
            "sampleRows": rows[:10],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.on_event("startup")
async def startup():
    print(f"\n  DataChat API Server running on port {API_PORT}")
    print(f"  Endpoints:")
    print(f"    POST /api/chat         — LLM chat with context")
    print(f"    POST /api/ingest/excel — Parse Excel files")
    print(f"    POST /api/ingest/pdf   — Parse PDF files")
    print(f"    POST /api/ingest/json  — Ingest JSON data")
    print(f"    GET  /health           — Health check\n")


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=API_PORT)
