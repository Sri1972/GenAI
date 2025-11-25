"""
Excel Service for NLP to Structured Data System

Handles Excel file operations including loading, parsing,
sheet management, and data extraction.
"""

import pandas as pd
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import logging
import aiofiles
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd
from utils.data_normalizer import DataNormalizer


class ExcelService:
    """
    Service for Excel file operations.
    
    Handles .xlsx and .xls files with support for multiple sheets,
    cell ranges, and Excel-specific formatting.
    """
    
    def __init__(self):
        self.logger = logging.getLogger("services.excel_service")
        self.logger.setLevel(logging.INFO)
        self.data_normalizer = DataNormalizer()
    
    async def load_excel(self, file_path: Union[str, Path], 
                        sheet_name: Optional[str] = None,
                        header_row: Optional[int] = 0,
                        cell_range: Optional[str] = None) -> pd.DataFrame:
        """
        Load Excel file into pandas DataFrame.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to load (None for first sheet)
            header_row: Row number to use as column headers (0-indexed)
            cell_range: Specific cell range to load (e.g., 'A1:D10')
            
        Returns:
            Pandas DataFrame with Excel data
        """
        try:
            path = Path(file_path)
            
            if not path.exists():
                raise FileNotFoundError(f"Excel file not found: {file_path}")
            
            # If no sheet name specified, get the first sheet name
            if sheet_name is None:
                sheet_names = await self.get_sheet_names(path)
                if sheet_names:
                    sheet_name = sheet_names[0]
                else:
                    sheet_name = 0  # Use first sheet by index
            
            # Load Excel file
            if cell_range:
                # Use openpyxl for cell range operations
                df = await self._load_cell_range(path, sheet_name, cell_range, header_row)
            else:
                # Use pandas for standard loading - ensure we get a DataFrame
                df = await asyncio.to_thread(
                    pd.read_excel,
                    path,
                    sheet_name=sheet_name,  # Now guaranteed to be a specific sheet
                    header=header_row,
                    engine='openpyxl' if path.suffix == '.xlsx' else 'xlrd'
                )
            
            # Ensure we have a DataFrame (not a dict)
            if isinstance(df, dict):
                # If somehow we still get a dict, take the first sheet
                first_key = list(df.keys())[0]
                df = df[first_key]
                self.logger.warning(f"Excel returned dict, using sheet: {first_key}")
            
            self.logger.info(f"Loaded Excel file: {file_path}, Shape: {df.shape}")
            return df
            
        except Exception as e:
            self.logger.error(f"Failed to load Excel file {file_path}: {str(e)}")
            raise
    
    async def _load_cell_range(self, file_path: Path, sheet_name: Optional[str], 
                              cell_range: str, header_row: Optional[int]) -> pd.DataFrame:
        """Load specific cell range from Excel file."""
        def _extract_range():
            workbook = load_workbook(file_path, read_only=True)
            
            # Select worksheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Extract cell range
            cells = worksheet[cell_range]
            
            # Convert to list of lists
            data = []
            for row in cells:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)
            
            workbook.close()
            return data
        
        # Run in thread to avoid blocking
        data = await asyncio.to_thread(_extract_range)
        
        # Convert to DataFrame
        if header_row is not None and len(data) > header_row:
            headers = data[header_row]
            data_rows = data[header_row + 1:]
            df = pd.DataFrame(data_rows, columns=headers)
        else:
            df = pd.DataFrame(data)
        
        return df
    
    async def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """
        Get list of sheet names in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        try:
            path = Path(file_path)
            
            def _get_sheets():
                if path.suffix == '.xlsx':
                    workbook = load_workbook(path, read_only=True)
                    sheets = workbook.sheetnames
                    workbook.close()
                    return sheets
                else:  # .xls
                    workbook = xlrd.open_workbook(path)
                    return workbook.sheet_names()
            
            sheets = await asyncio.to_thread(_get_sheets)
            return sheets
            
        except Exception as e:
            self.logger.error(f"Failed to get sheet names from {file_path}: {str(e)}")
            return []
    
    async def get_sheet_info(self, file_path: Union[str, Path]) -> Dict[str, Dict[str, Any]]:
        """
        Get information about all sheets in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with sheet information
        """
        try:
            path = Path(file_path)
            
            def _get_info():
                info = {}
                
                if path.suffix == '.xlsx':
                    workbook = load_workbook(path, read_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        # Get dimensions
                        max_row = worksheet.max_row
                        max_col = worksheet.max_column
                        
                        info[sheet_name] = {
                            "max_row": max_row,
                            "max_column": max_col,
                            "dimensions": f"{get_column_letter(max_col)}{max_row}",
                            "has_data": max_row > 0 and max_col > 0
                        }
                    
                    workbook.close()
                    
                else:  # .xls
                    workbook = xlrd.open_workbook(path)
                    
                    for sheet_name in workbook.sheet_names():
                        worksheet = workbook.sheet_by_name(sheet_name)
                        
                        info[sheet_name] = {
                            "max_row": worksheet.nrows,
                            "max_column": worksheet.ncols,
                            "dimensions": f"{get_column_letter(worksheet.ncols)}{worksheet.nrows}",
                            "has_data": worksheet.nrows > 0 and worksheet.ncols > 0
                        }
                
                return info
            
            return await asyncio.to_thread(_get_info)
            
        except Exception as e:
            self.logger.error(f"Failed to get sheet info from {file_path}: {str(e)}")
            return {}
    
    async def filter_dataframe(self, df: pd.DataFrame, 
                              filters: Dict[str, Any]) -> pd.DataFrame:
        """
        Filter DataFrame based on conditions.
        
        Args:
            df: Input DataFrame
            filters: Dictionary of column -> condition mappings
            
        Returns:
            Filtered DataFrame
        """
        try:
            filtered_df = df.copy()
            
            for column, condition in filters.items():
                if column not in filtered_df.columns:
                    continue
                
                if isinstance(condition, dict):
                    # Handle complex conditions
                    if 'operator' in condition and 'value' in condition:
                        op = condition['operator']
                        value = condition['value']
                        
                        if op == 'equals':
                            filtered_df = filtered_df[filtered_df[column] == value]
                        elif op == 'not_equals':
                            filtered_df = filtered_df[filtered_df[column] != value]
                        elif op == 'greater_than':
                            filtered_df = filtered_df[filtered_df[column] > value]
                        elif op == 'less_than':
                            filtered_df = filtered_df[filtered_df[column] < value]
                        elif op == 'contains':
                            filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(str(value), na=False)]
                        elif op == 'in':
                            filtered_df = filtered_df[filtered_df[column].isin(value)]
                else:
                    # Simple equality filter
                    filtered_df = filtered_df[filtered_df[column] == condition]
            
            return filtered_df
            
        except Exception as e:
            self.logger.error(f"Filtering failed: {str(e)}")
            return df
    
    async def aggregate_data(self, df: pd.DataFrame, 
                           group_by: List[str],
                           aggregations: Dict[str, str]) -> pd.DataFrame:
        """
        Perform aggregation operations on DataFrame.
        
        Args:
            df: Input DataFrame
            group_by: Columns to group by
            aggregations: Column -> aggregation function mappings
            
        Returns:
            Aggregated DataFrame
        """
        try:
            # Validate group_by columns
            valid_group_cols = [col for col in group_by if col in df.columns]
            if not valid_group_cols:
                return df
            
            # Prepare aggregation dict
            agg_dict = {}
            for column, agg_func in aggregations.items():
                if column in df.columns:
                    if agg_func in ['sum', 'mean', 'count', 'min', 'max', 'std']:
                        agg_dict[column] = agg_func
            
            if not agg_dict:
                return df.groupby(valid_group_cols).size().reset_index(name='count')
            
            # Perform aggregation
            result = df.groupby(valid_group_cols).agg(agg_dict).reset_index()
            
            # Flatten column names if needed
            if isinstance(result.columns, pd.MultiIndex):
                result.columns = ['_'.join(col).strip() if col[1] else col[0] 
                                for col in result.columns.values]
            
            return result
            
        except Exception as e:
            self.logger.error(f"Aggregation failed: {str(e)}")
            return df
    
    async def get_column_statistics(self, df: pd.DataFrame, 
                                   column: str) -> Dict[str, Any]:
        """
        Get statistical information about a column.
        
        Args:
            df: DataFrame to analyze
            column: Column name
            
        Returns:
            Dictionary with column statistics
        """
        try:
            if column not in df.columns:
                return {"error": f"Column '{column}' not found"}
            
            series = df[column]
            stats = {
                "name": column,
                "count": len(series),
                "non_null_count": series.count(),
                "null_count": series.isnull().sum(),
                "dtype": str(series.dtype),
                "unique_count": series.nunique()
            }
            
            # Add numeric statistics if applicable
            if pd.api.types.is_numeric_dtype(series):
                stats.update({
                    "mean": series.mean(),
                    "median": series.median(),
                    "std": series.std(),
                    "min": series.min(),
                    "max": series.max(),
                    "quantiles": {
                        "25%": series.quantile(0.25),
                        "50%": series.quantile(0.50),
                        "75%": series.quantile(0.75)
                    }
                })
            
            # Add text statistics for object columns
            elif series.dtype == 'object':
                stats.update({
                    "sample_values": series.dropna().head(5).tolist(),
                    "most_common": series.mode().tolist()[:3] if not series.mode().empty else []
                })
            
            return stats
            
        except Exception as e:
            self.logger.error(f"Failed to get column statistics: {str(e)}")
            return {"error": str(e)}
    
    async def search_data(self, df: pd.DataFrame, 
                         search_term: str, 
                         columns: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Search for term across DataFrame columns.
        
        Args:
            df: DataFrame to search
            search_term: Term to search for
            columns: Specific columns to search (None for all)
            
        Returns:
            DataFrame with matching rows
        """
        try:
            if columns:
                search_columns = [col for col in columns if col in df.columns]
            else:
                search_columns = df.select_dtypes(include=['object']).columns.tolist()
            
            if not search_columns:
                return pd.DataFrame()
            
            # Create search mask
            mask = pd.Series([False] * len(df))
            
            for column in search_columns:
                column_mask = df[column].astype(str).str.contains(
                    search_term, case=False, na=False
                )
                mask = mask | column_mask
            
            return df[mask]
            
        except Exception as e:
            self.logger.error(f"Search failed: {str(e)}")
            return pd.DataFrame()
    
    async def export_to_excel(self, df: pd.DataFrame, 
                             file_path: Union[str, Path],
                             sheet_name: str = "Sheet1") -> bool:
        """
        Export DataFrame to Excel file.
        
        Args:
            df: DataFrame to export
            file_path: Output file path
            sheet_name: Name of the sheet
            
        Returns:
            True if successful
        """
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            await asyncio.to_thread(
                df.to_excel,
                path,
                sheet_name=sheet_name,
                index=False
            )
            
            self.logger.info(f"Exported to Excel: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Export to Excel failed: {str(e)}")
            return False